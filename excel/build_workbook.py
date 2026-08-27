"""
Generates the Excel verification workbook: independently recomputes the
same IS 456 calculations as the app/ Python modules, using native Excel
formulas (not values copied from Python) - so opening this file and
comparing it against run_example.py's printed output is a genuine
cross-check, not just a display of the same numbers twice.

Run:
    python excel/build_workbook.py

Creates: excel/RCC_Beam_Design.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, size=9, color="666666")


def add_title(ws, text, span=3):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


# =============================================================================
# Sheet 1: Inputs
#   B3=Span, B4=Width, B5=OverallDepth, B6=Cover, B7=StirrupDia, B8=BarDia(trial),
#   B9=fck, B10=fy, B11=xu,max/d, B12=DL, B13=LL, B14=UnitWt,
#   B15=SelectedBarCount, B16=StirrupLegs
# =============================================================================
ws_in = wb.active
ws_in.title = "Inputs"
add_title(ws_in, "RCC BEAM DESIGN - INPUTS (B1)")
set_col_widths(ws_in, [32, 14, 30])

rows = [
    (3, "Span (mm)", 5000),
    (4, "Width, b (mm)", 300),
    (5, "Overall Depth, D (mm)", 500),
    (6, "Clear Cover (mm)", 25),
    (7, "Stirrup Diameter (mm)", 8),
    (8, "Main Bar Diameter, trial (mm)", 16),
    (9, "Concrete Grade, fck (N/mm2)", 25),
    (10, "Steel Grade, fy (N/mm2)", 415),
    (12, "Dead Load, excl. self-weight (kN/m)", 12),
    (13, "Live Load (kN/m)", 8),
    (14, "Unit weight of RCC (kN/m3)", 25),
    (15, "Selected Main Bar Count (for verification)", 4),
    (16, "Stirrup Legs", 2),
]
for r, label, value in rows:
    ws_in.cell(row=r, column=1, value=label).font = LABEL_FONT
    ws_in.cell(row=r, column=2, value=value)

ws_in.cell(row=11, column=1, value="xu,max / d ratio (Cl. 38.1)").font = LABEL_FONT
ws_in.cell(row=11, column=2,
           value='=IF(B10=250,0.53,IF(B10=415,0.48,IF(B10=500,0.46,IF(B10=550,0.44,"?"))))')

ws_in.cell(row=18, column=1, value="NOTE: uses the trial bar diameter for effective depth "
                                    "(not the optimizer's final selection - see Python tool "
                                    "for the actual bar-selection loop).").font = NOTE_FONT

# =============================================================================
# Sheet 2: Load Calculation
# =============================================================================
ws_load = wb.create_sheet("Load Calculation")
add_title(ws_load, "LOAD CALCULATION", span=3)
set_col_widths(ws_load, [32, 14, 35])

ws_load.cell(row=3, column=1, value="Effective depth, d (mm)").font = LABEL_FONT
ws_load.cell(row=3, column=2, value="=Inputs!B5-Inputs!B6-Inputs!B7-Inputs!B8/2")
ws_load.cell(row=3, column=3, value="D - cover - stirrup dia - bar dia/2 (Cl. 25.4)").font = NOTE_FONT

ws_load.cell(row=4, column=1, value="Self-weight (kN/m)").font = LABEL_FONT
ws_load.cell(row=4, column=2, value="=(Inputs!B4/1000)*(Inputs!B5/1000)*Inputs!B14")

ws_load.cell(row=5, column=1, value="Total service load (kN/m)").font = LABEL_FONT
ws_load.cell(row=5, column=2, value="=Inputs!B12+Inputs!B13+B4")

ws_load.cell(row=6, column=1, value="Factored load, wu (kN/m)").font = LABEL_FONT
ws_load.cell(row=6, column=2, value="=1.5*B5")
ws_load.cell(row=6, column=3, value="1.5(DL+LL) - Table 18 / Cl. 36.4.1").font = NOTE_FONT

ws_load.cell(row=7, column=1, value="Design Moment, Mu (kNm)").font = LABEL_FONT
ws_load.cell(row=7, column=2, value="=B6*(Inputs!B3/1000)^2/8")
ws_load.cell(row=7, column=3, value="wu*L^2/8, simply supported UDL").font = NOTE_FONT

ws_load.cell(row=8, column=1, value="Design Shear, Vu (kN)").font = LABEL_FONT
ws_load.cell(row=8, column=2, value="=B6*(Inputs!B3/1000)/2")
ws_load.cell(row=8, column=3, value="wu*L/2, simply supported UDL").font = NOTE_FONT

# =============================================================================
# Sheet 3: Flexural Design (singly reinforced)
# =============================================================================
ws_flex = wb.create_sheet("Flexural Design")
add_title(ws_flex, "FLEXURAL DESIGN - SINGLY REINFORCED", span=3)
set_col_widths(ws_flex, [32, 18, 40])

ws_flex.cell(row=3, column=1, value="Effective depth, d (mm)").font = LABEL_FONT
ws_flex.cell(row=3, column=2, value="='Load Calculation'!B3")

ws_flex.cell(row=4, column=1, value="Design Moment, Mu (kNm)").font = LABEL_FONT
ws_flex.cell(row=4, column=2, value="='Load Calculation'!B7")

ws_flex.cell(row=5, column=1, value="xu,max / d ratio").font = LABEL_FONT
ws_flex.cell(row=5, column=2, value="=Inputs!B11")

ws_flex.cell(row=6, column=1, value="Mu,lim (kNm)").font = LABEL_FONT
ws_flex.cell(row=6, column=2, value="=0.36*B5*(1-0.42*B5)*Inputs!B9*Inputs!B4*B3^2/1000000")
ws_flex.cell(row=6, column=3, value="IS 456 Annex G-1.1(c)").font = NOTE_FONT

ws_flex.cell(row=7, column=1, value="Doubly reinforced required?").font = LABEL_FONT
ws_flex.cell(row=7, column=2, value='=IF(B4>B6,"YES - see Python tool","NO")')

ws_flex.cell(row=8, column=1, value="Ast, required (mm^2)").font = LABEL_FONT
ws_flex.cell(row=8, column=2,
             value='=IF(B4<=B6, 0.5*(Inputs!B9/Inputs!B10)*Inputs!B4*B3*'
                   '(1-SQRT(1-4.6*B4*1000000/(Inputs!B9*Inputs!B4*B3^2))), "N/A")')
ws_flex.cell(row=8, column=3, value="Annex G-1.1(b)").font = NOTE_FONT

ws_flex.cell(row=9, column=1, value="Ast, min (mm^2)").font = LABEL_FONT
ws_flex.cell(row=9, column=2, value="=0.85*Inputs!B4*B3/Inputs!B10")
ws_flex.cell(row=9, column=3, value="Cl. 26.5.1.1(a)").font = NOTE_FONT

ws_flex.cell(row=10, column=1, value="Ast, max (mm^2)").font = LABEL_FONT
ws_flex.cell(row=10, column=2, value="=0.04*Inputs!B4*Inputs!B5")
ws_flex.cell(row=10, column=3, value="Cl. 26.5.1.1(b)").font = NOTE_FONT

ws_flex.cell(row=11, column=1, value="Ast, design (mm^2)").font = LABEL_FONT
ws_flex.cell(row=11, column=2, value='=IF(B7="NO",MAX(B8,B9),"N/A")')

ws_flex.cell(row=12, column=1, value="Flexure Status").font = LABEL_FONT
ws_flex.cell(row=12, column=2,
             value='=IF(B7="NO",IF(B11<=B10,"SAFE","FAIL - exceeds Ast,max"),"Doubly reinforced - use Python tool")')

# =============================================================================
# Sheet 4: Shear Design
# =============================================================================
ws_shear = wb.create_sheet("Shear Design")
add_title(ws_shear, "SHEAR DESIGN (VERTICAL STIRRUPS)", span=3)
set_col_widths(ws_shear, [32, 18, 35])

ws_shear.cell(row=3, column=1, value="Effective depth, d (mm)").font = LABEL_FONT
ws_shear.cell(row=3, column=2, value="='Load Calculation'!B3")

ws_shear.cell(row=4, column=1, value="Design Shear, Vu (kN)").font = LABEL_FONT
ws_shear.cell(row=4, column=2, value="='Load Calculation'!B8")

ws_shear.cell(row=5, column=1, value="tau_v (N/mm2)").font = LABEL_FONT
ws_shear.cell(row=5, column=2, value="=B4*1000/(Inputs!B4*B3)")
ws_shear.cell(row=5, column=3, value="Cl. 40.1").font = NOTE_FONT

ws_shear.cell(row=6, column=1, value="Ast provided (mm^2)").font = LABEL_FONT
ws_shear.cell(row=6, column=2, value="=Inputs!B15*PI()*Inputs!B8^2/4")

ws_shear.cell(row=7, column=1, value="pt, % tension steel provided").font = LABEL_FONT
ws_shear.cell(row=7, column=2, value="=100*B6/(Inputs!B4*B3)")

ws_shear.cell(row=8, column=1, value="tau_c (N/mm2)").font = LABEL_FONT
ws_shear.cell(row=8, column=2, value="=IF(C27=C26,B29,B29+(B7-C26)/(C27-C26)*(B30-B29))")
ws_shear.cell(row=8, column=3, value="Table 19 (interpolated below), Cl. 40.2.1").font = NOTE_FONT

ws_shear.cell(row=9, column=1, value="tau_c,max (N/mm2)").font = LABEL_FONT
ws_shear.cell(row=9, column=2, value="=INDEX(B47:G47,MATCH(Inputs!B9,B46:G46,0))")
ws_shear.cell(row=9, column=3, value="Table 20, Cl. 40.2.3").font = NOTE_FONT

ws_shear.cell(row=10, column=1, value="Section adequate?").font = LABEL_FONT
ws_shear.cell(row=10, column=2, value='=IF(B5<=B9,"YES","NO - increase b or D")')

ws_shear.cell(row=11, column=1, value="Reinforcement basis").font = LABEL_FONT
ws_shear.cell(row=11, column=2, value='=IF(B5<=B8,"Minimum (Cl. 40.3)","Designed (Cl. 40.4)")')

ws_shear.cell(row=12, column=1, value="Vus (kN)").font = LABEL_FONT
ws_shear.cell(row=12, column=2, value="=IF(B5>B8,B4-B8*Inputs!B4*B3/1000,0)")

ws_shear.cell(row=13, column=1, value="Asv, 2-legged (mm^2)").font = LABEL_FONT
ws_shear.cell(row=13, column=2, value="=Inputs!B16*PI()*Inputs!B7^2/4")

ws_shear.cell(row=14, column=1, value="Spacing, calculated (mm)").font = LABEL_FONT
ws_shear.cell(row=14, column=2,
              value='=IF(B11="Minimum (Cl. 40.3)",0.87*Inputs!B10*B13/(0.4*Inputs!B4),'
                    '0.87*Inputs!B10*B13*B3/(B12*1000))')

ws_shear.cell(row=15, column=1, value="Spacing, max allowed (mm)").font = LABEL_FONT
ws_shear.cell(row=15, column=2, value="=MIN(0.75*B3,300)")
ws_shear.cell(row=15, column=3, value="Cl. 26.5.1.5").font = NOTE_FONT

ws_shear.cell(row=16, column=1, value="Spacing, provided (mm)").font = LABEL_FONT
ws_shear.cell(row=16, column=2, value="=FLOOR(MIN(B14,B15),25)")

ws_shear.cell(row=17, column=1, value="Shear Status").font = LABEL_FONT
ws_shear.cell(row=17, column=2, value='=IF(AND(B10="YES",B16>0),"SAFE","FAIL")')

# --- Table 19 interpolation working area ---
ws_shear.cell(row=25, column=1, value="Working: Table 19 interpolation").font = NOTE_FONT
ws_shear.cell(row=26, column=1, value="Row index (lower pt)")
ws_shear.cell(row=26, column=2, value="=IFERROR(MATCH(B7,A32:A44,1),1)")
ws_shear.cell(row=26, column=3, value="=INDEX($A$32:$A$44,B26)")
ws_shear.cell(row=27, column=1, value="Row index (upper pt)")
ws_shear.cell(row=27, column=2, value="=IF(B26<13,B26+1,B26)")
ws_shear.cell(row=27, column=3, value="=INDEX($A$32:$A$44,B27)")
ws_shear.cell(row=28, column=1, value="Column index (fck)")
ws_shear.cell(row=28, column=2, value="=MATCH(Inputs!B9,B31:G31,0)+1")
ws_shear.cell(row=29, column=1, value="Lower tau_c")
ws_shear.cell(row=29, column=2, value="=INDEX($A$32:$G$44,B26,B28)")
ws_shear.cell(row=30, column=1, value="Upper tau_c")
ws_shear.cell(row=30, column=2, value="=INDEX($A$32:$G$44,B27,B28)")

# --- Table 19 data (IS 456) ---
ws_shear.cell(row=31, column=1, value="pt (%)")
for col, grade in zip("BCDEFG", [15, 20, 25, 30, 35, 40]):
    ws_shear[f"{col}31"] = grade
table19 = {
    0.15: [0.28, 0.28, 0.29, 0.29, 0.29, 0.30], 0.25: [0.35, 0.36, 0.36, 0.37, 0.37, 0.38],
    0.50: [0.46, 0.48, 0.49, 0.50, 0.50, 0.51], 0.75: [0.54, 0.56, 0.57, 0.59, 0.59, 0.60],
    1.00: [0.60, 0.62, 0.64, 0.66, 0.67, 0.68], 1.25: [0.64, 0.67, 0.70, 0.71, 0.73, 0.74],
    1.50: [0.68, 0.72, 0.74, 0.76, 0.78, 0.79], 1.75: [0.71, 0.75, 0.78, 0.80, 0.82, 0.84],
    2.00: [0.71, 0.79, 0.82, 0.84, 0.86, 0.88], 2.25: [0.71, 0.81, 0.85, 0.88, 0.90, 0.92],
    2.50: [0.71, 0.82, 0.88, 0.91, 0.93, 0.95], 2.75: [0.71, 0.82, 0.90, 0.94, 0.96, 0.98],
    3.00: [0.71, 0.82, 0.92, 0.96, 0.99, 1.01],
}
for i, (pt, values) in enumerate(table19.items()):
    r = 32 + i
    ws_shear.cell(row=r, column=1, value=pt)
    for col, val in zip("BCDEFG", values):
        ws_shear[f"{col}{r}"] = val

# --- Table 20 data (IS 456) - placed BELOW Table 19 (rows 32-44) to avoid overlap ---
ws_shear.cell(row=46, column=1, value="Grade")
ws_shear.cell(row=47, column=1, value="tau_c,max")
for col, grade, tcmax in zip("BCDEFG", [15, 20, 25, 30, 35, 40], [2.5, 2.8, 3.1, 3.5, 3.7, 4.0]):
    ws_shear[f"{col}46"] = grade
    ws_shear[f"{col}47"] = tcmax

# =============================================================================
# Sheet 5: Reinforcement
# =============================================================================
ws_reinf = wb.create_sheet("Reinforcement")
add_title(ws_reinf, "REINFORCEMENT DETAILING", span=3)
set_col_widths(ws_reinf, [32, 18, 35])

ws_reinf.cell(row=3, column=1, value="Selected bar diameter (mm)").font = LABEL_FONT
ws_reinf.cell(row=3, column=2, value="=Inputs!B8")
ws_reinf.cell(row=4, column=1, value="Selected bar count").font = LABEL_FONT
ws_reinf.cell(row=4, column=2, value="=Inputs!B15")
ws_reinf.cell(row=5, column=1, value="Ast provided (mm^2)").font = LABEL_FONT
ws_reinf.cell(row=5, column=2, value="=B4*PI()*B3^2/4")
ws_reinf.cell(row=6, column=1, value="Ast required (from Flexural Design)").font = LABEL_FONT
ws_reinf.cell(row=6, column=2, value="='Flexural Design'!B8")
ws_reinf.cell(row=7, column=1, value="Excess steel (%)").font = LABEL_FONT
ws_reinf.cell(row=7, column=2, value="=(B5-B6)/B6*100")
ws_reinf.cell(row=8, column=1, value="Clear bar spacing (mm)").font = LABEL_FONT
ws_reinf.cell(row=8, column=2, value="=(Inputs!B4-2*Inputs!B6-2*Inputs!B7-B4*B3)/(B4-1)")
ws_reinf.cell(row=8, column=3, value="Cl. 26.3.2").font = NOTE_FONT
ws_reinf.cell(row=9, column=1, value="Min required spacing (mm)").font = LABEL_FONT
ws_reinf.cell(row=9, column=2, value="=MAX(B3,25)")
ws_reinf.cell(row=10, column=1, value="Spacing check").font = LABEL_FONT
ws_reinf.cell(row=10, column=2, value='=IF(B8>=B9,"OK","FAIL - bars too close")')
ws_reinf.cell(row=11, column=1, value="Development length, Ld (mm)").font = LABEL_FONT
ws_reinf.cell(row=11, column=2, value="=B3*0.87*Inputs!B10/(4*INDEX(B22:F22,MATCH(Inputs!B9,B21:F21,0)))")
ws_reinf.cell(row=11, column=3, value="Cl. 26.2.1, 26.2.1.1").font = NOTE_FONT

ws_reinf.cell(row=20, column=1, value="Working: bond stress table (deformed bars)").font = NOTE_FONT
ws_reinf.cell(row=21, column=1, value="Grade")
ws_reinf.cell(row=22, column=1, value="tau_bd (deformed)")
for col, grade, tau_plain in zip("BCDEF", [20, 25, 30, 35, 40], [1.2, 1.4, 1.5, 1.7, 1.9]):
    ws_reinf[f"{col}21"] = grade
    ws_reinf[f"{col}22"] = round(tau_plain * 1.6, 4)

# =============================================================================
# Sheet 6: Serviceability
# =============================================================================
ws_serv = wb.create_sheet("Serviceability")
add_title(ws_serv, "SERVICEABILITY - DEFLECTION (SPAN/DEPTH)", span=3)
set_col_widths(ws_serv, [32, 18, 35])

ws_serv.cell(row=3, column=1, value="Effective depth, d (mm)").font = LABEL_FONT
ws_serv.cell(row=3, column=2, value="='Load Calculation'!B3")
ws_serv.cell(row=4, column=1, value="Span (mm)").font = LABEL_FONT
ws_serv.cell(row=4, column=2, value="=Inputs!B3")
ws_serv.cell(row=5, column=1, value="Basic L/d ratio (simply supported)").font = LABEL_FONT
ws_serv.cell(row=5, column=2, value="=IF(B4<=10000,20,20*10000/B4)")
ws_serv.cell(row=5, column=3, value="Cl. 23.2.1").font = NOTE_FONT
ws_serv.cell(row=6, column=1, value="Ast required").font = LABEL_FONT
ws_serv.cell(row=6, column=2, value="='Flexural Design'!B8")
ws_serv.cell(row=7, column=1, value="Ast provided").font = LABEL_FONT
ws_serv.cell(row=7, column=2, value="=Reinforcement!B5")
ws_serv.cell(row=8, column=1, value="fs, service stress (N/mm2)").font = LABEL_FONT
ws_serv.cell(row=8, column=2, value="=0.58*Inputs!B10*(B6/B7)")
ws_serv.cell(row=9, column=1, value="pt, % tension steel provided").font = LABEL_FONT
ws_serv.cell(row=9, column=2, value="=100*B7/(Inputs!B4*B3)")
ws_serv.cell(row=10, column=1, value="kt (modification factor)").font = LABEL_FONT
ws_serv.cell(row=10, column=2, value="=MAX(0.4,MIN(2,1/(0.225+0.00322*B8-0.625*LOG10(B9))))")
ws_serv.cell(row=10, column=3, value="Fig. 4 (SP:16 fit)").font = NOTE_FONT
ws_serv.cell(row=11, column=1, value="kc (compression steel factor)").font = LABEL_FONT
ws_serv.cell(row=11, column=2, value=1)
ws_serv.cell(row=11, column=3, value="1.0 - singly reinforced only in this sheet").font = NOTE_FONT
ws_serv.cell(row=12, column=1, value="Allowable L/d").font = LABEL_FONT
ws_serv.cell(row=12, column=2, value="=B5*B10*B11")
ws_serv.cell(row=13, column=1, value="Actual L/d").font = LABEL_FONT
ws_serv.cell(row=13, column=2, value="=B4/B3")
ws_serv.cell(row=14, column=1, value="Serviceability Status").font = LABEL_FONT
ws_serv.cell(row=14, column=2, value='=IF(B13<=B12,"SAFE","FAIL - increase depth or reduce span")')

# =============================================================================
# Sheet 7: Design Summary (dashboard)
# =============================================================================
ws_sum = wb.create_sheet("Design Summary")
add_title(ws_sum, "RCC BEAM DESIGN SUMMARY", span=3)
set_col_widths(ws_sum, [32, 30, 10])

ws_sum.cell(row=3, column=1, value="Span (m)").font = LABEL_FONT
ws_sum.cell(row=3, column=2, value="=Inputs!B3/1000")
ws_sum.cell(row=4, column=1, value="Width x Overall Depth (mm)").font = LABEL_FONT
ws_sum.cell(row=4, column=2, value='=Inputs!B4&" x "&Inputs!B5')
ws_sum.cell(row=5, column=1, value="Concrete / Steel").font = LABEL_FONT
ws_sum.cell(row=5, column=2, value='="M"&Inputs!B9&" / Fe"&Inputs!B10')
ws_sum.cell(row=6, column=1, value="Factored Load (kN/m)").font = LABEL_FONT
ws_sum.cell(row=6, column=2, value="='Load Calculation'!B6")
ws_sum.cell(row=7, column=1, value="Design Moment, Mu (kNm)").font = LABEL_FONT
ws_sum.cell(row=7, column=2, value="='Load Calculation'!B7")
ws_sum.cell(row=8, column=1, value="Design Shear, Vu (kN)").font = LABEL_FONT
ws_sum.cell(row=8, column=2, value="='Load Calculation'!B8")
ws_sum.cell(row=9, column=1, value="Ast, required (mm^2)").font = LABEL_FONT
ws_sum.cell(row=9, column=2, value="='Flexural Design'!B8")
ws_sum.cell(row=10, column=1, value="Ast, provided (mm^2)").font = LABEL_FONT
ws_sum.cell(row=10, column=2, value="=Reinforcement!B5")
ws_sum.cell(row=11, column=1, value="Main Reinforcement").font = LABEL_FONT
ws_sum.cell(row=11, column=2, value='=Inputs!B15&" x "&Inputs!B8&"mm"')
ws_sum.cell(row=12, column=1, value="Stirrups").font = LABEL_FONT
ws_sum.cell(row=12, column=2, value='=Inputs!B16&"-legged "&Inputs!B7&"mm @ "&\'Shear Design\'!B16&"mm c/c"')
ws_sum.cell(row=13, column=1, value="Development Length, Ld (mm)").font = LABEL_FONT
ws_sum.cell(row=13, column=2, value="=Reinforcement!B11")

ws_sum.cell(row=15, column=1, value="Flexure Status").font = LABEL_FONT
ws_sum.cell(row=15, column=2, value="='Flexural Design'!B12")
ws_sum.cell(row=16, column=1, value="Shear Status").font = LABEL_FONT
ws_sum.cell(row=16, column=2, value="='Shear Design'!B17")
ws_sum.cell(row=17, column=1, value="Serviceability Status").font = LABEL_FONT
ws_sum.cell(row=17, column=2, value="=Serviceability!B14")
ws_sum.cell(row=18, column=1, value="OVERALL STATUS").font = Font(bold=True, size=12)
ws_sum.cell(row=18, column=2, value='=IF(AND(B15="SAFE",B16="SAFE",B17="SAFE"),"SAFE","NOT SAFE - CHECK ABOVE")')

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ws_sum.conditional_formatting.add("B15:B18", CellIsRule(operator="equal", formula=['"SAFE"'], fill=green_fill))
ws_sum.conditional_formatting.add("B15:B18", CellIsRule(operator="notEqual", formula=['"SAFE"'], fill=red_fill))

wb.save("excel/RCC_Beam_Design.xlsx")
print("Saved excel/RCC_Beam_Design.xlsx")